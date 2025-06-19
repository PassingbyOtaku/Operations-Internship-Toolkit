import os
import requests
from urllib.parse import urlparse, unquote
from openpyxl import load_workbook
from openpyxl.comments import Comment

xlsx_path = r'文件路径'
img_root = os.path.join(os.path.dirname(xlsx_path), 'img')#默认导出到与xlsx同级目录下的img文件夹

wb = load_workbook(xlsx_path)
ws = wb.active

# 获取表头
header = [cell.value for cell in ws[1]]
col_idx = {
    '序号': header.index('序号') + 1,
    #列名、表头需要根据实际修改
    'O': header.index('前30天的直播数据（需完整展示直播时长、累计观看、弹幕数）') + 1,
    'P': header.index('所有直播场次的数据截图（如下图所示，需提供活动期间内所有直播场次的数据截图）') + 1,
}

for row in ws.iter_rows(min_row=2):
    idx = str(row[col_idx['序号']-1].value)
    for col, col_letter in [('O', 'O'), ('P', 'P')]:
        cell = row[col_idx[col]-1]
        urls = str(cell.value).replace('，', ',').split(',')
        folder = os.path.join(img_root, idx)
        os.makedirs(folder, exist_ok=True)
        img_names = []
        for i, url in enumerate(urls):
            url = url.strip()
            if not url or url.lower() == '(空)':
                continue
            try:
                ext = os.path.splitext(unquote(urlparse(url).path))[1]
                if not ext or len(ext) > 6:
                    ext = '.jpg'
                fname = f"{idx}_{col_letter}_{i+1}{ext}"
                fpath = os.path.join(folder, fname)
                img_names.append(fname)
                if not os.path.exists(fpath):
                    r = requests.get(url, timeout=15)
                    with open(fpath, 'wb') as imgf:
                        imgf.write(r.content)
            except Exception as e:
                print(f"下载失败: {url} 错误: {e}")
        # 写批注
        if img_names:
            comment = Comment('\n'.join(img_names), "Copilot")
            cell.comment = comment

wb.save(xlsx_path)
print("图片下载和批注插入完成。")